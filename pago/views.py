from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from itertools import groupby
from .models import Pago, Caja
from .forms import PagoForm, CajaForm
from pedidos.models import Pedido


def _solo_cajero_admin(request):
    """Retorna True si el usuario NO tiene permiso (es mesero u otro rol no autorizado)."""
    return not (request.user.rol in ('ADMIN', 'CAJERO') or request.user.is_superuser)


@login_required
def pago_dashboard(request):
    if _solo_cajero_admin(request):
        return render(request, 'usuarios/login.html', {'vista': 'sin_permisos'})

    form_apertura = CajaForm()
    form = PagoForm()

    caja_activa = Caja.objects.filter(estado='ABIERTA').first()

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'abrir_caja':
            if caja_activa:
                messages.error(
                    request,
                    f'Ya hay una caja abierta (Caja #{caja_activa.pk} — {caja_activa.cajero}. '
                    f'Debes cerrarla antes de abrir una nueva.'
                )
                return redirect('pago:dashboard')
            form_apertura = CajaForm(request.POST)
            if form_apertura.is_valid():
                apertura = form_apertura.save(commit=False)
                apertura.estado = 'ABIERTA'
                apertura.save()
                messages.success(request, ' Caja abierta correctamente.')
                return redirect('pago:dashboard')
            else:
                messages.error(request, ' Revisa los campos e intenta de nuevo.')

        elif action == 'cerrar_caja':
            caja_id = request.POST.get('caja_id')
            try:
                caja = Caja.objects.get(pk=caja_id, estado='ABIERTA')
                caja.estado = 'CERRADA'
                caja.fecha_cierre = timezone.now()
                caja.save()
                messages.success(request, ' Caja cerrada correctamente.')
            except Caja.DoesNotExist:
                messages.error(request, ' No se encontró la caja o ya está cerrada.')
            return redirect('pago:dashboard')

        elif action == 'editar_caja':
            caja_id       = request.POST.get('caja_id')
            observaciones = request.POST.get('observaciones', '').strip()
            try:
                caja = Caja.objects.get(pk=caja_id)
                caja.observaciones = observaciones
                caja.save()
                return JsonResponse({
                    'ok': True,
                    'observaciones': caja.observaciones or '—',
                })
            except Exception as e:
                return JsonResponse({'ok': False, 'error': str(e)}, status=400)

        else:
            post_data = request.POST.copy()
            pedido_id  = post_data.get('pedido')
            if pedido_id:
                try:
                    pedido = Pedido.objects.get(pk=pedido_id)
                    post_data['monto'] = pedido.total
                except Pedido.DoesNotExist:
                    pass
            form = PagoForm(post_data)
            if form.is_valid():
                pago = form.save(commit=False)
                if caja_activa:
                    pago.caja = caja_activa
                    pago.save()
                    pago.pedido.estado = 'PAGADO'
                    pago.pedido.save()
                    messages.success(request, 'Pedido registrado correctamente.', extra_tags='modal-pago')
                else:
                    messages.error(request, ' No puedes registrar pagos sin una caja abierta.')
                return redirect('pago:dashboard')

    ordenes_sin_pago = Pedido.objects.exclude(estado='PAGADO').order_by('-fecha_creacion')
    pagos_qs = Pago.objects.select_related('pedido').order_by('-fecha_pago')

    pagos_por_fecha = []
    for fecha, grupo in groupby(pagos_qs, key=lambda p: p.fecha_pago.date()):
        items = list(grupo)
        pagos_por_fecha.append({
            'fecha': fecha,
            'pagos': items,
            'total': sum(p.monto for p in items),
            'count': len(items),
        })

    context = {
        'form':             form,
        'form_apertura':    form_apertura,
        'pagos_por_fecha':  pagos_por_fecha,
        'ordenes_sin_pago': ordenes_sin_pago,
        'total_pagos':      pagos_qs.count(),
        'pagos_aprobados':  pagos_qs.count(),
        'pagos_pendientes': 0,
        'monto_total':      pagos_qs.aggregate(t=Sum('monto'))['t'] or 0,
        'nombre':           request.user.get_full_name() or request.user.username,
        'cajas':            Caja.objects.select_related('cajero').all().order_by('fecha_apertura'),
        'tab_activo':       request.GET.get('tab', 'pendientes'),
        'caja_activa':      caja_activa,
    }
    return render(request, 'pago/dashboard.html', context)


@login_required
def pago_editar(request, pk):
    if _solo_cajero_admin(request):
        return render(request, 'usuarios/login.html', {'vista': 'sin_permisos'})

    pago = get_object_or_404(Pago, pk=pk)
    form = PagoForm(request.POST or None, instance=pago)
    if form.is_valid():
        form.save()
        messages.success(request, ' Pago actualizado.')
        return redirect('pago:dashboard')
    context = {
        'form':   form,
        'pago':   pago,
        'nombre': request.user.get_full_name() or request.user.username,
    }
    return render(request, 'pago/form.html', context)


@login_required
def pago_eliminar(request, pk):
    if _solo_cajero_admin(request):
        return render(request, 'usuarios/login.html', {'vista': 'sin_permisos'})

    pago = get_object_or_404(Pago, pk=pk)
    if request.method == 'POST':
        pago.delete()
        messages.success(request, ' Pago eliminado.')
    return redirect('pago:dashboard')


@login_required
def caja_detalle(request, pk):
    if _solo_cajero_admin(request):
        return render(request, 'usuarios/login.html', {'vista': 'sin_permisos'})

    caja_seleccionada = get_object_or_404(Caja, pk=pk)
    pagos_caja = Pago.objects.filter(caja=caja_seleccionada).select_related('pedido').order_by('-fecha_pago')

    total_ingresos   = pagos_caja.aggregate(t=Sum('monto'))['t'] or 0
    total_pendientes = 0

    form_apertura = CajaForm()
    form = PagoForm()

    ordenes_sin_pago = Pedido.objects.exclude(estado='PAGADO').order_by('-fecha_creacion')
    pagos_qs = Pago.objects.select_related('pedido').order_by('-fecha_pago')

    pagos_por_fecha = []
    for fecha, grupo in groupby(pagos_qs, key=lambda p: p.fecha_pago.date()):
        items = list(grupo)
        pagos_por_fecha.append({
            'fecha': fecha,
            'pagos': items,
            'total': sum(p.monto for p in items),
            'count': len(items),
        })

    context = {
        'form':               form,
        'form_apertura':      form_apertura,
        'pagos_por_fecha':    pagos_por_fecha,
        'ordenes_sin_pago':   ordenes_sin_pago,
        'total_pagos':        pagos_qs.count(),
        'pagos_aprobados':    pagos_qs.count(),
        'pagos_pendientes':   0,
        'monto_total':        pagos_qs.aggregate(t=Sum('monto'))['t'] or 0,
        'nombre':             request.user.get_full_name() or request.user.username,
        'cajas':              Caja.objects.select_related('cajero').all().order_by('fecha_apertura'),
        'caja_seleccionada':  caja_seleccionada,
        'pagos_caja':         pagos_caja,
        'total_ingresos':     total_ingresos,
        'total_pendientes':   total_pendientes,
        'saldo_final':        caja_seleccionada.monto_inicial + total_ingresos,
        'tab_activo':         'detalle-caja',
        'caja_activa':        Caja.objects.filter(estado='ABIERTA').first(),
    }
    return render(request, 'pago/dashboard.html', context)


# ── REPORTES ────────────────────────────────────────────────────────

def _pagos_reporte_queryset():
    return Pago.objects.select_related('pedido', 'pedido__cliente').order_by('-fecha_pago')


@login_required
def pagos_exportar_excel(request):
    if _solo_cajero_admin(request):
        return render(request, 'usuarios/login.html', {'vista': 'sin_permisos'})

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    pagos = _pagos_reporte_queryset()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagos registrados'

    encabezados = ['Orden', 'Cliente', 'Método', 'Monto', 'Referencia', 'Estado', 'Fecha', 'Hora']
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        celda = ws.cell(row=1, column=col)
        celda.font      = Font(bold=True, color='F5ECD7')
        celda.fill      = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        celda.alignment = Alignment(horizontal='center')

    for pg in pagos:
        ws.append([
            pg.pedido.numero_orden if pg.pedido else '—',
            pg.pedido.cliente.nombre_completo if pg.pedido and pg.pedido.cliente else '—',
            pg.get_metodo_pago_display(),
            float(pg.monto),
            pg.referencia or '—',
            'Pagado',
            pg.fecha_pago.strftime('%d/%m/%Y'),
            pg.fecha_pago.strftime('%I:%M %p'),
        ])

    for i, ancho in enumerate([14, 26, 16, 12, 20, 12, 14, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pagos_registrados.xlsx"'
    wb.save(response)
    return response


@login_required
def pagos_exportar_pdf(request):
    if _solo_cajero_admin(request):
        return render(request, 'usuarios/login.html', {'vista': 'sin_permisos'})

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    pagos = _pagos_reporte_queryset()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pagos_registrados.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=landscape(letter),
        topMargin=1.2*cm, bottomMargin=1.2*cm,
        leftMargin=1.2*cm, rightMargin=1.2*cm,
    )
    estilos   = getSampleStyleSheet()
    elementos = [
        Paragraph('Porras Asadero — Pagos registrados', estilos['Title']),
        Paragraph(f'Generado el {timezone.now().strftime("%d/%m/%Y %I:%M %p")}', estilos['Normal']),
        Spacer(1, 0.5*cm),
    ]

    datos = [['Orden', 'Cliente', 'Método', 'Monto', 'Referencia', 'Estado', 'Fecha', 'Hora']]
    total = 0
    for pg in pagos:
        total += pg.monto
        datos.append([
            pg.pedido.numero_orden if pg.pedido else '—',
            pg.pedido.cliente.nombre_completo if pg.pedido and pg.pedido.cliente else '—',
            pg.get_metodo_pago_display(),
            f'${pg.monto:,.0f}',
            pg.referencia or '—',
            'Pagado',
            pg.fecha_pago.strftime('%d/%m/%Y'),
            pg.fecha_pago.strftime('%I:%M %p'),
        ])
    datos.append(['', '', '', f'${total:,.0f}', '', '', '', ''])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0),  (-1, 0),  colors.HexColor('#C0392B')),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.HexColor('#F5ECD7')),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, -1), 8.5),
        ('ALIGN',         (3, 1),  (3, -1),  'RIGHT'),
        ('GRID',          (0, 0),  (-1, -1), 0.5, colors.HexColor('#D4C4A0')),
        ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.HexColor('#FDF7EC'), colors.HexColor('#F5ECD7')]),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND',    (0, -1), (-1, -1), colors.HexColor('#F5ECD7')),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    return response


@login_required
def pagos_imprimir(request):
    if _solo_cajero_admin(request):
        return render(request, 'usuarios/login.html', {'vista': 'sin_permisos'})

    pagos = _pagos_reporte_queryset()
    return render(request, 'pago/pagos_imprimir.html', {
        'pagos':  pagos,
        'ahora':  timezone.now(),
        'total':  pagos.aggregate(t=Sum('monto'))['t'] or 0,
    })